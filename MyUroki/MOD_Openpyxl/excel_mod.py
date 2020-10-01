import openpyxl, os, pprint
xl= openpyxl.load_workbook(os.path.dirname(__file__)+'/'+"200.xlsx")

print("\nВывод имен всех имеющихся листов книги:", xl.sheetnames)

print("\nВывод активного листа:",xl.active)
xl.active=0
print("\nВывод измененного активного листа:",xl.active)

act_sh=xl['Главный']

print("\nЗначение ячейки:", act_sh['I2'].value, 
      "\nНомера строки и колонки:", act_sh['I2'].row,"/", act_sh['I2'].column,
      "\nКоордината ячейки:",act_sh['I2'].coordinate)

print("\nВывод значения ячейки через метод Cell:", act_sh.cell(row=2, column=9).value )

print(f'''\nВывод размера листа-{act_sh.title}.
  Всего колонок:''', act_sh.max_column,
  "\n  Всего строк:", act_sh.max_row)

for i in range(1, act_sh.max_column):
    print('Индекс колонки-', openpyxl.utils.get_column_letter(i), f' ={i}')

print('Возвращает индекс колонки по ее буквенному значению:', 
       openpyxl.utils.column_index_from_string("AA"),
       openpyxl.utils.column_index_from_string("KN"))

list= []
for i_cikl_po_stroke in act_sh['A1':'H30']:
    list_0=[]
    for y_cikl_po_stolbcu in i_cikl_po_stroke:
        list_0.append(y_cikl_po_stolbcu.value)
    list.append(list_0)
pprint.pprint(list)

#print(act_sh['B'])
print(act_sh[5])
xl.close()

xl_new= openpyxl.Workbook()
xl_new.active.title="Import data"
xl_new.create_sheet(title="New Sheet")
xl_new.create_sheet(title="New Sheet1")
xl_new.create_sheet(title="New Sheet2")
xl_new.remove_sheet(xl_new["New Sheet1"])
print(xl_new.active)
file_save=xl_new.active

for i in range(len(list)):
    for y in range(len(list[i])):
        file_save.cell(row=i+1, column=y+1).value = list[i][y]

from openpyxl.styles import Font
file_save["A1"].font = Font(size=24, italic=True)
file_save.row_dimensions[1].height=70
file_save.column_dimensions["B"].width=20

file_save.merge_cells("I3:L5")

file_save.freeze_panes="A2"

xl_new.save(os.path.dirname(__file__)+'/'+"300.xlsx")
        
