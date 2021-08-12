import openpyxl
import os

cur_dir=f"{os.path.dirname(__file__)}"
user_dir="user_data"
user_file_name="abonent.xlsx"
file_db=f"{cur_dir}\{user_dir}\{user_file_name}"

def opxls():
    if os.path.isfile(file_db):  
        xl=openpyxl.load_workbook(file_db)
        xl.active=0
        activ_sh= xl.active
        list_user=[]
        for i_str in range(1, activ_sh.max_row+1):
            if i_str==1: 
                list_0=[]
            dict_0={}
            for y_column in range(1, activ_sh.max_column+1):
            
                if type(activ_sh.cell(row=i_str, column=y_column).value) is float:
                    activ_sh.cell(row=i_str, column=y_column).value= round(activ_sh.cell(row=i_str, column=y_column).value, 2)
            
                if i_str==1:
                    list_0.append(activ_sh.cell(row=i_str, column=y_column).value)
                if i_str>1:
                    dict_0[list_0[y_column-1]]= activ_sh.cell(row=i_str, column=y_column).value
                    if list_0[y_column-1]=="tel1" or list_0[y_column-1]=="tel2":
                        dict_0[list_0[y_column-1]]= "".join(str(dict_0[list_0[y_column-1]]).split("-"))

            list_user.append(dict_0)
        #print(list_user)
        xl.close
        return list_user
    return None

def find_user(key, find_val, key2="tel2"):
    list_dict=opxls()
    result_find=[]
    if list_dict!=None:
        for i in list_dict:
            if i.get(key)==find_val:
                result_find.append(i)
            if i.get(key2)==find_val:
                result_find.append(i)
        return result_find
            


#print(find_user("tel1", "380999476550"))



#print(list_user)
#print(activ_sh.max_row, activ_sh.max_column)